"""2025/3/22版本：
1.cUrl命令.xlsx中新增请求头字段，以请求方式开头，然后是接口路径，然后是请求头信息 √
2.需要从请求头字段里，获取请求方式以及content-type/Accept √
3.从正式的response里，获取所有单层的键值对，另外从多层键值对里匹配*id的key，如果没有则随机获取3个键值对作为断言 √
"""

import re
import json
import logging
from urllib.parse import urlparse
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from collections import OrderedDict
from response_process import ResponseProcess
from urllib.parse import unquote


logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("log.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# # 登录接口固定header
# LOGIN_HEADER_FIXED = ''

response_process = ResponseProcess()

# ---------------------- 新增解码函数 ----------------------
def decode_url_encoded(value):
    """安全解码URL编码值"""
    try:
        decoded_value = unquote(value)
        logger.debug(f"解码URL编码值: {value} -> {decoded_value}")
        return decoded_value
    except Exception as e:
        logger.error(f"解码失败：{value}，错误：{str(e)}", exc_info=True)
        return value

def parse_query_with_decode(query_str):
    """解析并解码查询参数"""
    logger.debug(f"开始解析查询参数: {query_str}")
    decoded_query = []
    if query_str:
        pairs = query_str.split('&')
        logger.debug(f"解析到 {len(pairs)} 个查询参数对")
        for pair in pairs:
            if '=' in pair:
                key, val = pair.split('=', 1)
                decoded_key = decode_url_encoded(key)
                decoded_val = decode_url_encoded(val)
                decoded_query.append(f"{decoded_key}={decoded_val}")
                logger.debug(f"解析查询参数对: {key}={val} -> {decoded_key}={decoded_val}")
            else:
                decoded_part = decode_url_encoded(pair)
                decoded_query.append(decoded_part)
                logger.debug(f"解析独立查询参数: {pair} -> {decoded_part}")
    result = '@@'.join(decoded_query)
    logger.debug(f"最终解析的查询参数: {result}")
    return result

def decode_nested_values(data):
    """递归解码嵌套结构中的URL编码值"""
    logger.debug("开始解码嵌套结构数据")
    if isinstance(data, dict):
        decoded_dict = {}
        for k, v in data.items():
            decoded_key = decode_url_encoded(k)
            logger.debug(f"处理字典键: {k} -> {decoded_key}")
            decoded_value = decode_nested_values(v)
            decoded_dict[decoded_key] = decoded_value
        return decoded_dict
    elif isinstance(data, list):
        decoded_list = []
        for i, item in enumerate(data):
            logger.debug(f"处理列表第 {i} 项")
            decoded_list.append(decode_nested_values(item))
        return decoded_list
    elif isinstance(data, str):
        decoded_str = decode_url_encoded(data)
        logger.debug(f"解码字符串: {data} -> {decoded_str}")
        return decoded_str
    else:
        logger.debug(f"无需解码的数据类型: {type(data)}")
        return data

# ---------------------- 核心解析逻辑 ----------------------
def parse_curl(curl_command):
    """从cURL命令解析协议、路径、查询参数和请求体"""
    logger.info(f"开始解析cURL命令，长度: {len(curl_command)} 字符")
    try:
        # URL解析
        url_match = re.search(r"'(https?://[^']+)'", curl_command) or re.search(r'"(https?://[^"]+)"', curl_command)
        if not url_match:
            logger.warning("未找到有效的URL")
            url = ''
        else:
            url = url_match.group(1)
            logger.info(f"解析到原始URL: {url}")

        parsed_url = urlparse(url) if url else urlparse('')
        logger.debug(f"解析URL结果: scheme={parsed_url.scheme}, path={parsed_url.path}, query={parsed_url.query}")

        # 请求体解析
        data_match = re.search(r"--data-raw '(.*?)'", curl_command) or re.search(r"-d '(.*?)'", curl_command)
        raw_data = data_match.group(1) if data_match else ''
        logger.debug(f"原始请求体数据: {raw_data[:100]}...（共{len(raw_data)}字节）")

        # 查询参数解码
        decoded_query = parse_query_with_decode(parsed_url.query) if parsed_url.query else ''
        logger.info(f"解码后的查询参数: {decoded_query}")

        # 请求体解码处理
        decoded_data = raw_data
        try:
            content_type = 'application/json' if 'application/json' in curl_command.lower() else ''
            if content_type and raw_data:
                logger.debug("开始解析JSON请求体")
                json_data = json.loads(raw_data)
                decoded_json = decode_nested_values(json_data)
                decoded_data = json.dumps(decoded_json, ensure_ascii=False)
                logger.info(f"解码后的JSON请求体: {decoded_data[:200]}...")
            elif 'application/x-www-form-urlencoded' in curl_command.lower() and raw_data:
                logger.debug("处理表单编码数据")
                decoded_pairs = []
                pairs = raw_data.split('&')
                for pair in pairs:
                    if '=' in pair:
                        k, v = pair.split('=', 1)
                        dk, dv = unquote(k), unquote(v)
                        decoded_pairs.append(f"{dk}={dv}")
                        logger.debug(f"表单参数解码: {k}={v} -> {dk}={dv}")
                    else:
                        decoded = unquote(pair)
                        decoded_pairs.append(decoded)
                        logger.debug(f"独立表单参数解码: {pair} -> {decoded}")
                decoded_data = '&'.join(decoded_pairs)
                logger.info(f"解码后的表单数据: {decoded_data[:200]}...")
        except Exception as e:
            logger.error(f"请求体解码过程中发生错误: {str(e)}", exc_info=True)
            decoded_data = raw_data  # 出错时保留原始数据

        return {
            'protocol': parsed_url.scheme,
            'path': parsed_url.path,
            'query': decoded_query,
            'data': decoded_data,
            'content_type': content_type
        }
    except Exception as e:
        logger.error("解析cURL命令时发生未预期的错误", exc_info=True)
        return {
            'protocol': '',
            'path': '',
            'query': '',
            'data': '',
            'content_type': ''
        }

def parse_content_type(header_str):
    """解析Content-Type或Accept头部"""
    logger.debug(f"开始解析请求头内容: {header_str[:100]}...")
    lines = [line.strip() for line in header_str.split('\n') if line.strip()]
    logger.debug(f"清理后头部行数: {len(lines)}")

    content_type = None
    accept = None

    for line in lines:
        if ':' not in line:
            logger.debug(f"跳过非头部行: {line}")
            continue

        key, value = line.split(':', 1)
        key = key.strip().lower()
        value = value.strip()

        if key == 'content-type':
            content_type = value.split(';')[0].strip()
            logger.info(f"解析到Content-Type: {content_type}")
            break  # 优先使用Content-Type
        elif key == 'accept' and not content_type:
            accept = value.split(';')[0].strip().split(',')[0]
            logger.info(f"回退到Accept头: {accept}")

    logger.info(f"最终内容类型判断: {content_type or accept or '未识别'}")
    return content_type or accept

def parse_request_headers(header_str):
    """解析请求头字段获取请求方法和内容类型"""
    logger.debug(f"解析请求头字符串: {header_str[:100]}...")
    method = None
    header_lines = [line.strip() for line in header_str.split('\n') if line.strip()]

    # 解析请求方法
    first_line = header_lines[0] if header_lines else ''
    method_match = re.match(r'(GET|POST|PUT|DELETE|PATCH|HEAD|OPTIONS)\s+', first_line)
    if method_match:
        method = method_match.group(1).lower()
        logger.info(f"识别到请求方法: {method.upper()}")
    else:
        logger.warning(f"无法识别请求方法，首行内容: {first_line}")

    # 解析内容类型
    content_type = parse_content_type(header_str)
    return {
        'method': method,
        'headers': content_type
    }

# ---------------------- 数据处理逻辑 ----------------------
CONTENT_TYPE_MAP = {
    "application/json": 1,
    "application/x-www-form-urlencoded": 2,
    "multipart/form-data": 3,
    "text/plain": 4,
    "text/xml": 5,
    "tcbs一代客户端": 8,
    "application/x-amf": 9
}

def read_input_excel(input_file):
    """读取输入Excel文件"""
    logger.info(f"开始读取输入文件: {input_file}")
    try:
        wb = load_workbook(input_file)
        ws = wb.active
        data = {}
        logger.info(f"工作表'{ws.title}'中共有{ws.max_row}行数据")

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            try:
                if len(row) >= 5 and row[0].value and row[1].value:
                    interface_num = str(row[0].value).strip()
                    interface_name = str(row[1].value).strip()
                    curl_command = str(row[2].value).strip() if row[2].value else ''
                    request_headers = str(row[3].value).strip() if row[3].value else ''
                    response = str(row[4].value).strip() if row[4].value else ''

                    logger.debug(f"读取第{row_idx}行: \n"
                                f"接口编号={interface_num}, 名称={interface_name}\n "
                                f"cUrl：\n{curl_command}\n "
                                f"请求头：\n{request_headers}\n"
                                f"响应：\n{response}")

                    data[interface_name] = (
                        interface_num,
                        curl_command,
                        request_headers,
                        response
                    )
            except Exception as e:
                logger.error(f"处理第{row_idx}行时发生错误", exc_info=True)
                continue

        logger.info(f"成功读取{len(data)}个有效接口配置")
        return data

    except FileNotFoundError:
        logger.error(f"文件不存在: {input_file}", exc_info=True)
        raise
    except Exception as e:
        logger.error("读取Excel文件时发生未预期错误", exc_info=True)
        raise

def generate_interface_data(interface_name, parsed_data, interface_num, headers):
    """生成接口文档数据行"""
    logger.info(f"开始生成接口文档数据: {interface_name}")
    try:
        # 确定header值
        if interface_name == "总行管理员登录":
            header_value = None
            logger.debug("特殊处理登录接口header")
        else:
            header_value = 'Authorization=123'
            logger.debug("设置默认Authorization头")

        # 解析内容类型
        content_type = parse_request_headers(headers)['headers']
        encoding_type = CONTENT_TYPE_MAP.get(content_type, 7)
        logger.info(f"内容类型映射: {content_type} -> {encoding_type}")

        # 解析请求方法
        method_info = parse_request_headers(headers)
        method = method_info['method'] or 'get'
        logger.info(f"确定请求方法: {method}")

        return {
            '接口编号': interface_num,
            '项目名称': '请填写',
            '接口名称': interface_name,
            '接口协议(http、https)': parsed_data.get('protocol', 'http'),
            '接口路径': parsed_data.get('path', ''),
            '接口请求方法(get、post、put、delete)': method,
            '请求体编码类型': encoding_type,
            'header': header_value,
            'query-params': parsed_data.get('query', ''),
            'body': parsed_data.get('data', '')
        }
    except Exception as e:
        logger.error(f"生成接口数据失败: {interface_name}", exc_info=True)
        raise

def process_body_data(parsed_data):
    """处理body数据，根据内容类型转换为键值对格式"""
    logger.debug("开始处理请求体数据")
    try:
        content_type = parsed_data.get('content_type', '')
        data = parsed_data.get('data', '')
        logger.info(f"原始请求体内容类型: {content_type}, 数据长度: {len(data)}")

        if 'application/json' in content_type.lower() and data:
            logger.debug("处理JSON格式请求体")
            try:
                json_data = json.loads(data)
                items = []
                for key, value in json_data.items():
                    if isinstance(value, (list, dict)):
                        processed_value = json.dumps(value, ensure_ascii=False)
                        logger.debug(f"处理复杂类型字段: {key}=[...]")
                    else:
                        processed_value = str(value)
                        logger.debug(f"处理简单字段: {key}={value}")
                    items.append(f"{key}={processed_value}")
                result = '@@'.join(items)
                logger.info(f"JSON请求体处理后结果: {result[:200]}...")
                return result
            except json.JSONDecodeError:
                logger.warning("JSON解析失败，保持原始数据")
                return data
        else:
            processed = data.replace('&', '@@') if data else ''
            logger.info(f"非JSON请求体处理后结果: {processed[:200]}...")
            return processed
    except Exception as e:
        logger.error("处理请求体数据时发生错误", exc_info=True)
        return parsed_data.get('data', '')

def generate_testcase_rows(interface_name, parsed_data, response, case_num):
    """生成测试用例的多行数据"""
    logger.info(f"开始生成测试用例: {interface_name} 用例{case_num}")
    try:
        # Header处理
        if interface_name == "总行管理员登录":
            header_value = None
            logger.debug("登录接口不设置Authorization头")
        else:
            header_value = 'Authorization=Bearer #{总行管理员登录_RECMSG}_token#'
            logger.debug("设置动态Authorization头")

        # 处理请求体
        processed_body = process_body_data(parsed_data)
        logger.debug(f"处理后的请求体参数: {processed_body[:100]}...")

        # 基础用例数据
        base_data = OrderedDict([
            ('用例状态', 1),
            ('用例编号', f"{interface_name}_{case_num:02d}"),
            ('中间件类型', 'HTTP'),
            ('其他配置项', ''),
            ('接口名称', interface_name),
            ('用例名称', f"{interface_name}_{case_num:02d}"),
            ('前置动作', ''),
            ('header', header_value),
            ('接口路径', ''),
            ('query-params', parsed_data.get('query', '')),
            ('body', processed_body),
            ('校验动作', ''),
            ('校验项', ''),
            ('断言方式', ''),
            ('预期值', ''),
            ('后置动作', '')
        ])
        logger.debug(f"基础用例数据: {dict(base_data)}")

        # 断言处理
        assertions = []
        if response:
            logger.info(f"开始处理响应数据，长度: {len(response)}")
            try:
                processed_responses = response_process.load_json(response)
                logger.info(f"从响应中提取到{len(processed_responses)}个断言项")

                for item in processed_responses:
                    for key, value in item.items():
                        assertions.append({
                            '校验项': key,
                            '断言方式': "等于",
                            '预期值': value
                        })
                        logger.debug(f"添加断言: {key} 等于 {value}")
            except Exception as e:
                logger.error("处理响应生成断言时出错", exc_info=True)
                assertions.append({
                    '校验项': "status_code",
                    '断言方式': "等于",
                    '预期值': 200
                })
        else:
            logger.warning("无响应数据，使用默认状态码断言")
            assertions.append({
                '校验项': "status_code",
                '断言方式': "等于",
                '预期值': 200
            })

        # 构建多行数据
        rows = []
        for idx, assertion in enumerate(assertions, 1):
            row = base_data.copy()
            row.update(assertion)
            rows.append(row)
            logger.debug(f"生成第{idx}行用例数据: {assertion}")

        # 合并列配置
        merge_cols = list(range(11))  # 合并前11列
        logger.info(f"将合并{len(merge_cols)}列，生成{len(rows)}行数据")

        return rows, len(rows), merge_cols
    except Exception as e:
        logger.error(f"生成测试用例行失败: {interface_name}", exc_info=True)
        raise


def main(input_file="./cUrl命令.xlsx"):
    """主函数入口"""
    logger.info("========== 开始执行主流程 ==========")
    logger.info(f"输入文件路径: {input_file}")

    try:
        # ==================== 读取输入文件 ====================
        logger.info(">> 阶段1/4 正在读取输入文件...")
        try:
            curl_data = read_input_excel(input_file)
            logger.info(f"[OK] 成功读取 {len(curl_data)} 个接口配置")
        except FileNotFoundError:
            logger.error(f"× 文件不存在: {input_file}")
            print(f"错误：输入文件 {input_file} 不存在")
            return
        except Exception as e:
            logger.error("× 读取Excel文件发生致命错误", exc_info=True)
            print(f"读取Excel失败：{str(e)}")
            return

        # ==================== 初始化工作簿 ====================
        logger.info(">> 阶段2/4 正在初始化输出文档...")
        interface_wb = Workbook()
        testcase_wb = Workbook()

        # 删除默认Sheet
        for wb in [interface_wb, testcase_wb]:
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
                logger.debug(f"已删除默认工作表: {wb.sheetnames}")

        # ==================== 样式配置 ====================
        logger.debug("正在配置样式...")
        header_font = Font(name='宋体', size=12, bold=True)
        header_fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
        header_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center', text_rotation=0)
        column_width = 40

        # ==================== 接口文档表 ====================
        interface_sheet = interface_wb.create_sheet("接口文档")
        logger.debug("已创建接口文档工作表")

        interface_headers = [
            "接口编号", "项目名称", "接口名称", "接口协议(http、https)",
            "接口路径", "接口请求方法(get、post、put、delete)",
            '请求体编码类型(1:"application/json", 2:"application/x-www-form-urlencoded", 3:"multipart/form-data", 4:"text/plain", 5:"text/xml", 7:"none", 8:"TCBS一代客户端", 9:"application/x-amf")',
            "header(http请求头)", "query-params(http请求参数，以@@分割拼接在url后)", "body(http请求体)"
        ]
        interface_sheet.append(interface_headers)
        logger.debug("已写入接口文档表头")

        # 设置列宽和样式
        for col in range(1, len(interface_headers) + 1):
            col_letter = get_column_letter(col)
            interface_sheet.column_dimensions[col_letter].width = column_width
        for cell in interface_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        logger.debug("接口文档样式配置完成")

        # ==================== 测试用例表 ====================
        testcase_sheet = testcase_wb.create_sheet("测试用例")
        logger.debug("已创建测试用例工作表")

        testcase_headers = [
            "用例状态(0=不执行@@1=发送+接收@@2=仅接收)", "用例编号", "中间件类型", "其他配置项", "接口名称",
            "用例名称", "前置动作", "header(http请求头参数替换项)", "接口路径(接口路径参数替换项)",
            "query-params(http请求参数替换项)", "body(http请求体参数替换项)", "校验动作", "校验项",
            "断言方式", "预期值", "后置动作"
        ]
        testcase_sheet.append(testcase_headers)
        logger.debug("已写入测试用例表头")

        # 设置列宽和样式
        for col in range(1, len(testcase_headers) + 1):
            col_letter = get_column_letter(col)
            testcase_sheet.column_dimensions[col_letter].width = column_width
        for cell in testcase_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        logger.debug("测试用例样式配置完成")

        # ==================== 处理接口数据 ====================
        logger.info(">> 阶段3/4 正在处理接口数据...")
        total_interfaces = 0
        total_testcases = 0
        current_row = 2

        for interface_name, (interface_num, curl_command, headers, response) in curl_data.items():
            try:
                logger.info(f"┌── 开始处理接口: [{interface_num}] {interface_name}")
                logger.debug(f"原始cURL命令长度: {len(curl_command)}")

                # 解析cURL命令
                parsed = parse_curl(curl_command)
                logger.debug(f"解析结果: 协议={parsed['protocol']} 路径={parsed['path']}")

                # 生成接口文档行
                interface_row = generate_interface_data(interface_name, parsed, interface_num, headers)
                interface_sheet.append(list(interface_row.values()))
                total_interfaces += 1
                logger.debug(f"接口文档已写入，当前总数: {total_interfaces}")

                # 生成测试用例
                case_num = 1
                rows, row_count, merge_cols = generate_testcase_rows(interface_name, parsed, response, case_num)

                # 写入测试用例
                start_row = current_row
                for row in rows:
                    testcase_sheet.append(list(row.values()))
                    current_row += 1
                total_testcases += len(rows)
                logger.debug(f"已写入{len(rows)}条测试用例")

                # 合并单元格
                end_row = start_row + row_count - 1
                for col_idx in merge_cols:
                    col_letter = get_column_letter(col_idx + 1)
                    testcase_sheet.merge_cells(f'{col_letter}{start_row}:{col_letter}{end_row}')
                    # 设置垂直居中
                    for r in range(start_row, end_row + 1):
                        testcase_sheet[f'{col_letter}{r}'].alignment = Alignment(vertical='center')
                logger.debug(f"合并单元格范围: {start_row}-{end_row}行")

                logger.info(f"└── 成功处理: 生成{len(rows)}条用例")
                print(f"成功处理接口：{interface_name}")

            except Exception as e:
                logger.error(f"处理接口异常: {interface_name}", exc_info=True)
                print(f"处理接口 [{interface_name}] 时出错：{str(e)}")
                continue

        logger.info(f"[OK] 完成接口处理 总数: {total_interfaces} 接口, 生成: {total_testcases} 测试用例")

        # ==================== 保存文件 ====================
        logger.info(">> 阶段4/4 正在保存文件...")
        try:
            interface_output = "接口文档.xlsx"
            testcase_output = "测试用例.xlsx"

            interface_wb.save(interface_output)
            testcase_wb.save(testcase_output)

            logger.info(f"文件保存成功\n接口文档: {interface_output}\n测试用例: {testcase_output}")
            print("成功生成文件：\n- 接口文档.xlsx\n- 测试用例.xlsx")

        except PermissionError:
            logger.error("文件保存失败: Excel文件被占用")
            print("错误：请关闭正在使用的Excel文件后重试")
        except Exception as e:
            logger.error("保存文件时发生未知错误", exc_info=True)

    except Exception as e:
        logger.critical("主流程发生未预期异常", exc_info=True)
    finally:
        logger.info("========== 流程执行结束 ==========\n")


if __name__ == "__main__":
    main()